
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from typing import List
from pathlib import Path
import io, re
import pandas as pd
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

app = FastAPI(title="Comparador Cotizaciones V5 Industrial")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def clean(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()

def norm(s):
    s = clean(s).lower()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n"),("²","2")]:
        s = s.replace(a, b)
    return s

def parse_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^0-9,.\-]", "", str(v or "").strip())
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    try:
        return float(s)
    except Exception:
        return None

def pdf_pages(data):
    import pdfplumber
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        return [p.extract_text(x_tolerance=1, y_tolerance=3) or "" for p in pdf.pages]

def detect_currency(text):
    t = text.upper()
    if "USD" in t or "U$S" in t or "U$" in t or "DOLAR" in t or "DÓLAR" in t:
        return "USD"
    if "$" in t or "PESOS" in t or "ARS" in t:
        return "ARS"
    return "USD"

def detect_provider(filename, text):
    l = text.lower()
    if "provemet" in l or "complemet" in l:
        return "Provemet / Complemet"
    if "ivanar" in l:
        return "IVANAR"
    if "la tornillera" in l:
        return "La Tornillera"
    if "hg confecciones" in l or "ropa de trabajo" in l:
        return "HG Confecciones"
    if "ingeniería boggio" in l or "ingenieria boggio" in l:
        return "Ingeniería Boggio"
    if "marlew" in l:
        return "Marlew"
    if "ateco cables" in l:
        return "Ateco"

    lines = [clean(x) for x in text.splitlines() if clean(x)]
    for line in lines[:18]:
        nl = norm(line)
        if any(x in nl for x in ["presupuesto","cotizacion","fecha","cliente","cuit","domicilio","telefono"]):
            continue
        if 4 <= len(line) <= 65:
            return line[:65]
    return Path(filename).stem

def detect_quote(filename, text):
    patterns = [
        r"COTIZACI[ÓO]N\s*N[°º]?\s*([0-9]+)",
        r"N[°º]\s*:\s*([0-9]+)",
        r"PRESUPUESTO\s*Nro\.\s*([0-9\-]+)",
        r"PRESUPUESTO\s+NUMERO\s*:\s*([0-9]+)",
        r"PRESUPUESTO\s+([0-9]+)",
        r"Presupuesto:\s*([0-9]+)",
        r"N[úu]mero:\s*([A-Z0-9\-]+)",
        r"PR\s*([0-9]+)"
    ]
    for p in patterns:
        m = re.search(p, text, re.I | re.S)
        if m:
            return clean(m.group(1))
    return Path(filename).stem

def normalize_group(text):
    t = clean(text)
    nt = norm(t).upper()

    m = re.search(r"(\d+\s*[xX]\s*\d+(?:[\.,]\d+)?(?:\s*\+\s*B?\d+)?\s*(?:mm2|MM2|mm²|MM²)?)", t)
    if m:
        form = re.sub(r"\s+", "", m.group(1).replace("X","x").replace(",",".").replace("MM2","mm²").replace("mm2","mm²"))
        flags = []
        ln = norm(t)
        if "blind" in ln or "pantalla" in ln or "malla" in ln:
            flags.append("BLIND")
        if "verde" in ln and "amarillo" in ln:
            flags.append("VA")
        if "pvc" in ln:
            flags.append("PVC")
        return "_".join(["CABLE", form] + flags)

    maps = {
        "EC 0210":"CABLE_2x1mm²_BLIND","EC 0215":"CABLE_2x1.5mm²_BLIND",
        "EC 0307":"CABLE_3x0.75mm²_BLIND","EC 0410":"CABLE_4x1mm²_BLIND",
        "NF 11500":"CABLE_1x150mm²","OF 1210":"CABLE_12x1mm²",
        "NF 0215":"CABLE_2x1.5mm²","NF 0225":"CABLE_2x2.5mm²",
        "NF 0315":"CABLE_3x1.5mm²","NF 0325":"CABLE_3x2.5mm²",
        "NF 0425":"CABLE_4x2.5mm²","NF 0440":"CABLE_4x4mm²",
        "OF 0715":"CABLE_7x1.5mm²","VK 1160":"CABLE_1x16mm²_VA",
        "VK 0125":"CABLE_1x2.5mm²_VA",
    }
    for k, v in maps.items():
        if k in nt:
            return v

    if any(x in nt for x in ["TUBO","CANO","CAÑO"]):
        size = ""; sch = ""; diam = ""
        msize = re.search(r"(\d+\s*1/2|\d+\s*1/4|\d+\s*3/4|\d+/\d+|\d+)\s*''?", t)
        if msize:
            size = re.sub(r"\s+", "", msize.group(1))
        msch = re.search(r"SCH\s*(\d+)", t, re.I)
        if msch:
            sch = "SCH" + msch.group(1)
        mdiam = re.search(r"Ø\s*([\d\.,]+)", t)
        if mdiam:
            diam = "D" + mdiam.group(1).replace(",", ".")
        parts = ["TUBO"]
        if size: parts.append(size)
        if sch: parts.append(sch)
        if diam: parts.append(diam)
        return "_".join(parts) if len(parts) > 1 else clean(t)[:60].upper()

    u = nt
    for key in ["VARILLA ROSCADA", "TUERCA HEXAG", "ARANDELA LISA", "PERFIL UPN"]:
        if key in u:
            base = re.sub(r"\s+", " ", t.upper()).strip()
            base = re.sub(r"\s+0[,\.]00%.*$", "", base)
            return base[:80]

    if any(w in norm(t) for w in ["camisa", "pantalon", "pantalón", "remera", "buzo", "logo bordado"]):
        base = t.upper()
        base = re.sub(r"\b(GRIS|TOPO|HOMOLOGAD[AO]|OZ|M/L|M/CORTAS|1ºCALIDAD|1RA CALIDAD)\b", "", base, flags=re.I)
        base = re.sub(r"\s+", " ", base).strip()
        return base[:80]

    base = norm(t).upper()
    base = re.sub(r"[^A-Z0-9X/\"\. ]", " ", base)
    base = re.sub(r"\s+", " ", base).strip()
    return base[:80]

def make_item(filename, prov, cot, nro, codigo, marca, desc, unidad, cant, punit, subtotal, moneda="USD", iva_pct=21, notas="", minimo="", entrega="", parser=""):
    desc = clean(desc)
    grupo = normalize_group(desc if desc else codigo)
    subtotal = subtotal if subtotal is not None else (cant or 0) * (punit or 0)
    iva = subtotal * iva_pct / 100
    return {
        "archivo": filename, "proveedor": prov, "cotizacion": cot, "nro_item": str(nro),
        "codigo": codigo or "", "codigo_interno": "", "marca": marca or "", "descripcion": desc,
        "formacion": grupo, "grupo_comparable": grupo, "moneda": moneda, "unidad": unidad or "u",
        "cantidad_pedida": cant or 0, "cantidad_real": cant or 0, "precio_unitario": punit or 0,
        "subtotal_sin_iva": subtotal or 0, "iva_pct": iva_pct, "iva_monto": iva,
        "total_con_iva": subtotal + iva, "minimo_compra": minimo, "venta_fraccionada": "",
        "entrega": entrega, "notas": notas, "parser": parser, "validado": False
    }

def parse_provemet(filename, pages):
    text = "\n".join(pages)
    cot = detect_quote(filename, text)
    out = []
    pattern = re.compile(r"(TUBO\s+S/C\s+ASTM.+?BW)\s+(\d+)\s+([\d\.,]+)\s+([\d\.,]+)\s+Mt\s+([\d\.,]+)\s+([\d\.,]+)\s+Mt\s+([\d\.,]+)", re.I | re.S)
    for m in pattern.finditer(text):
        desc = clean(m.group(1))
        nro = int(m.group(2))
        cant = parse_num(m.group(4)) or 0
        punit = parse_num(m.group(6)) or 0
        subtotal = parse_num(m.group(7)) or cant * punit
        out.append(make_item(filename, "Provemet / Complemet", cot, nro, "", "PROVEMET", desc, "Mt", cant, punit, subtotal, "USD", 21, "Precio no incluye IVA", "", "En stock / sujeto a venta", "provemet_v5"))
    return out

def parse_ivanar(filename, pages):
    text = "\n".join(pages); cot = detect_quote(filename, text); out = []
    for line in [clean(x) for x in text.splitlines() if clean(x)]:
        m = re.match(r"^\s*([0-9]{4,8})\s+(.+?)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+(UN|KG|MT|M|U)\b", line, re.I)
        if not m: continue
        codigo, desc, cant_raw, punit_raw, subtotal_raw, unidad = m.groups()
        if any(x in norm(desc) for x in ["subtotal","total","iva","documento"]): continue
        cant = parse_num(cant_raw) or 1; punit = parse_num(punit_raw) or 0; subtotal = parse_num(subtotal_raw) or cant*punit
        out.append(make_item(filename, "IVANAR", cot, len(out)+1, codigo, "IVANAR", desc, unidad, cant, punit, subtotal, "ARS", 21, "", "No informado", "A confirmar", "ivanar_v5"))
    return out

def parse_tornillera(filename, pages):
    text = "\n".join(pages); cot = detect_quote(filename, text); out = []
    for line in [clean(x) for x in text.splitlines() if clean(x)]:
        m = re.match(r"^([\d\.,]+)\s+\*\s+(.+?)\s+[\d\.,]+%\s+([\d\.,]+)\s+([\d\.,]+)$", line, re.I)
        if not m: continue
        cant_raw, desc, punit_raw, subtotal_raw = m.groups()
        cant = parse_num(cant_raw) or 1; punit = parse_num(punit_raw) or 0; subtotal = parse_num(subtotal_raw) or cant*punit
        out.append(make_item(filename, "La Tornillera", cot, len(out)+1, "", "LA TORNILLERA", desc, "u", cant, punit, subtotal, "ARS", 21, "Material sujeto a disponibilidad", "No informado", "24/04/2026", "tornillera_v5"))
    return out

def parse_hg(filename, pages):
    text = "\n".join(pages); cot = detect_quote(filename, text); out = []
    for line in [clean(x) for x in text.splitlines() if clean(x)]:
        m = re.match(r"^(.+?)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)$", line)
        if not m: continue
        desc, cant_raw, punit_raw, subtotal_raw = m.groups()
        if not any(x in norm(desc) for x in ["camisa","pantalon","pantalón","remera","buzo","logo"]): continue
        cant = parse_num(cant_raw) or 1; punit = parse_num(punit_raw) or 0; subtotal = parse_num(subtotal_raw) or cant*punit
        out.append(make_item(filename, "HG Confecciones", cot, len(out)+1, "", "HG", desc, "u", cant, punit, subtotal, "ARS", 21, "Entrega inicia al recibir talles/seña", "No informado", "A confirmar", "hg_v5"))
    return out

def parse_boggio(filename, pages):
    text = "\n".join(pages); cot = detect_quote(filename, text); entrega = "15/05/2026" if "Fecha de Entrega: 15/05/2026" in text else "Consultar"
    lines = [clean(x) for x in text.splitlines() if clean(x)]
    blocks = []; cur = []; start = re.compile(r"^\d+\s+\d{3,6}\s+\[[^\]]+\]\s+", re.I)
    ignore = ("X ","Doc. no","Ingeniería Boggio","Remedios","Argentina -","www.","Cliente:","L De La Torre","Cond. IVA","CUIT:","Comercial:","Vencimiento:","Términos","Referencia","% de","N° Imagen","Subtotal USD","IVA 21%","Total USD","JUANI","PRECIOS NETOS","NETO PAGO","Su pedido","Para más","https:","Fecha de Entrega","Los precios","Las garantias","(*)Confirmar","Ud. fue","Página:")
    for ln in lines:
        if ln.startswith(ignore): continue
        if start.match(ln):
            if cur: blocks.append(" ".join(cur))
            cur = [ln]
        elif cur: cur.append(ln)
    if cur: blocks.append(" ".join(cur))
    out = []; brand_re = r"\b(INDECA|WENTINCK|FONSECA|IMSA|PRYSMIAN)\b"
    for b in blocks:
        cm = re.match(r"^(\d+)\s+(\d{3,6})\s+\[([^\]]+)\]\s+(.+)$", b, re.I)
        if not cm: continue
        nro, img, cod, rest = cm.groups()
        m = re.search(brand_re + r"\s*(?:\(\*\))?\s+([\d\.,]+)\s*(m|mt|mts|Unidad|unidad|un|u)?\s+([\d\.,]+)\s+IVA\s*21%\s+([\d\.,]+)", rest, re.I)
        if not m: continue
        marca = m.group(1).upper(); cant = parse_num(m.group(2)) or 1; unidad = "m" if clean(m.group(3)).lower().startswith("m") else "u"; punit = parse_num(m.group(4)) or 0; subtotal = parse_num(m.group(5)) or cant*punit
        desc = clean(rest[:m.start()]).replace("()", "").strip()
        out.append(make_item(filename, "Ingeniería Boggio", cot, nro, cod, marca, desc, unidad, cant, punit, subtotal, "USD", 21, "Confirmar stock" if "(*)" in b else "", "No informado", entrega, "boggio_v5"))
    return out

def parse_marlew(filename, pages):
    text = "\n".join(pages); cot = detect_quote(filename, text); lines = [clean(x) for x in text.splitlines() if clean(x)]
    starts = [i for i,l in enumerate(lines) if re.match(r"^\d+\s+\d{2,5}\s+MT\s+C[oó]digo:", l, re.I)]
    blocks = [" ".join(lines[s: starts[j+1] if j+1 < len(starts) else len(lines)]) for j,s in enumerate(starts)]
    out = []
    for b in blocks:
        m = re.search(r"^(\d+)\s+(\d{2,5})\s+MT\s+C[oó]digo:\s*(.*?)\s+Formaci[oó]n:\s*([^\s]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", b, re.I)
        if not m: continue
        nro, cant_raw, cod, form, punit_raw, net, total_raw = m.groups()
        cant = parse_num(cant_raw) or 1; punit = parse_num(punit_raw) or 0; subtotal = parse_num(total_raw) or cant*punit
        low = norm(b); entrega = "6/8 semanas" if "6/8 semanas" in low else "A confirmar"; minimo = ""; notas = []
        if "minimo de provision" in low: minimo = f"Mín. {int(cant)} m"
        if "unica bobina" in low or "no fraccionable" in low: minimo = f"Bobina única {int(cant)} m"; notas.append("No fraccionable")
        if "material en stock" in low: notas.append("En stock salvo venta")
        out.append(make_item(filename, "Marlew", cot, nro, cod, "MARLEW", f"{cod} {form}", "m", cant, punit, subtotal, "USD", 21, "; ".join(notas), minimo, entrega, "marlew_v5"))
    return out

def parse_ateco(filename, pages):
    text = "\n".join(pages); cot = detect_quote(filename, text); entrega = "5 días" if "Plazo de entrega: 5 dias" in text else "Consultar"; out = []
    for ln in [clean(x) for x in text.splitlines() if clean(x)]:
        m = re.match(r"^([0-9]{4}-[0-9]-[0-9]{6}-[0-9])\s+(\d+)\s+(.+?)\s+sa\s+([\d\.,]+)\s+([\d\.,]+)$", ln, re.I)
        if not m: continue
        cod, cant_raw, desc, punit_raw, sub_raw = m.groups()
        cant = parse_num(cant_raw) or 1; punit = parse_num(punit_raw) or 0; subtotal = parse_num(sub_raw) or cant*punit
        out.append(make_item(filename, "Ateco", cot, len(out)+1, cod, "ATECO", desc, "m", cant, punit, subtotal, "USD", 21, "Contado anticipo", "No informado", entrega, "ateco_v5"))
    return out

BAD_WORDS = ["cuit","iva responsable","domicilio","telefono","tel.","email","mail","www","cliente","fecha","presupuesto","cotizacion","cotización","validez","condicion","condición","forma de pago","subtotal","total","observacion","observación","pagina","página","banco","cbu","vendedor","factura","documento no valido","son pesos","son dolares"]
TECH_WORDS = ["cable","caño","cano","tubo","válvula","valvula","brida","codo","tee","reducción","reduccion","cupla","niple","curva","acople","chapa","perfil","aislación","aislacion","mm","pulg","sch","ansi","astm","inox","acero","cobre","pvc","xlpe","blindado","pantalla","motor","bomba","sensor","presostato","termómetro","manómetro","tablero","borne","interruptor","contactor","disyuntor","varilla","tuerca","arandela","camisa","pantalon","pantalón","remera","buzo","logo"]

def looks_like_item(line):
    l = norm(line)
    if len(line) < 12: return False
    if any(b in l for b in BAD_WORDS): return False
    nums = re.findall(r"\d+(?:[\.,]\d+)?", line)
    if len(nums) < 2: return False
    return any(w in l for w in TECH_WORDS)

def parse_universal(filename, pages):
    text = "\n".join(pages); prov = detect_provider(filename, text); cot = detect_quote(filename, text); moneda = detect_currency(text)
    out = []
    for line in [clean(x) for x in text.splitlines() if clean(x)]:
        if not looks_like_item(line): continue
        nums = [parse_num(x) for x in re.findall(r"\d+(?:[\.,]\d+)?", line)]
        nums = [x for x in nums if x is not None]
        if len(nums) < 2: continue
        subtotal = nums[-1]; punit = nums[-2]; cant = 1
        if punit and punit > 0:
            est = subtotal / punit
            if 0 < est < 100000: cant = round(est, 4)
        qmatch = re.match(r"^\s*([\d\.,]+)\s*(m|mt|mts|un|u|unidad|kg)?\b", line, re.I)
        if qmatch:
            q = parse_num(qmatch.group(1))
            if q and q > 0: cant = q
        unidad = "m" if re.search(r"\b(m|mt|mts|metro|metros)\b", line, re.I) else "u"
        mcode = re.search(r"\b([A-Z]{1,5}[- ]?\d{2,8}|[0-9]{3,8}[-/][0-9A-Z\-]+)\b", line, re.I)
        codigo = clean(mcode.group(1)) if mcode else ""
        out.append(make_item(filename, prov, cot, len(out)+1, codigo, "", line, unidad, cant, punit, subtotal, moneda, 21, "Detectado por lector universal: revisar", "A confirmar", "A confirmar", "universal_v5"))
    return out

def similarity(a, b):
    return SequenceMatcher(None, norm(a), norm(b)).ratio()

def auto_enrich_groups(items):
    for i,a in enumerate(items):
        if not a.get("grupo_comparable"): a["grupo_comparable"] = normalize_group(a.get("descripcion",""))
        for j,b in enumerate(items):
            if j <= i or a.get("proveedor") == b.get("proveedor"): continue
            ga, gb = a.get("grupo_comparable",""), b.get("grupo_comparable","")
            if ga and gb and ga == gb: continue
            if similarity(ga or a.get("descripcion",""), gb or b.get("descripcion","")) > 0.84:
                group = ga if len(ga) >= len(gb) else gb
                a["grupo_comparable"] = group; b["grupo_comparable"] = group
    return items

def compare_items(items):
    items = auto_enrich_groups(items)
    df = pd.DataFrame(items)
    if df.empty: return [], "No hay ítems para comparar."
    df["grupo_comparable"] = df["grupo_comparable"].fillna("").astype(str).str.strip()
    df = df[df["grupo_comparable"] != ""]
    comps, summary = [], []
    for group, g in df.groupby("grupo_comparable"):
        if g["proveedor"].astype(str).str.lower().nunique() < 2: continue
        note = "Monedas distintas: validar tipo de cambio." if g["moneda"].astype(str).nunique() > 1 else ""
        g = g.sort_values("precio_unitario"); best = g.iloc[0]; offers = []
        for _,r in g.iterrows():
            d = r.to_dict(); d["dif_unit_vs_mejor"] = float(r["precio_unitario"] - best["precio_unitario"]); d["dif_total_vs_mejor"] = float(r["total_con_iva"] - best["total_con_iva"]); d["recomendado_precio"] = float(r["precio_unitario"]) == float(best["precio_unitario"]); offers.append(d)
        comps.append({"grupo_comparable": group, "mejor_proveedor": best["proveedor"], "moneda": best["moneda"], "mejor_precio_unitario": float(best["precio_unitario"]), "mejor_total_con_iva": float(best["total_con_iva"]), "nota": note, "ofertas": offers})
        summary.append(f"{group}: mejor precio unitario {best['proveedor']} — {best['moneda']} {best['precio_unitario']:.2f}. {note} Validar equivalencia técnica, mínimos y entrega.")
    return comps, "\n".join(summary) if summary else "No hay grupos comparables con 2 o más proveedores. Revisá/normalizá la columna 'Grupo comparable'."

def build_excel(items, comps, summary):
    wb = Workbook(); thin = Side(style="thin", color="D9D9D9"); fill = PatternFill("solid", fgColor="1F4E78"); font = Font(color="FFFFFF", bold=True); green = PatternFill("solid", fgColor="C6EFCE"); yellow = PatternFill("solid", fgColor="FFF2CC")
    ws0 = wb.active; ws0.title = "Resumen Ejecutivo"; ws0["A1"] = "Resumen Ejecutivo - Comparativa de Cotizaciones"; ws0["A1"].font = Font(size=16, bold=True); ws0["A3"] = summary or "Sin resumen."; ws0["A3"].alignment = Alignment(wrap_text=True, vertical="top"); ws0.column_dimensions["A"].width = 120
    ws0["A6"]="Cantidad de ítems detectados"; ws0["B6"]=len(items); ws0["A7"]="Comparaciones encontradas"; ws0["B7"]=len(comps); ws0["A8"]="Proveedores detectados"; ws0["B8"]=len(set([i.get("proveedor","") for i in items]))
    ws = wb.create_sheet("Items Detectados")
    headers = ["Proveedor","Cotización","N°","Código","Marca","Descripción","Grupo comparable","Moneda","Cant. pedida","Cant. real","Unidad","P. unit.","Subtotal s/IVA","IVA %","IVA monto","Total c/IVA","Mínimo","Entrega","Notas","Parser"]
    for c,h in enumerate(headers,1):
        cell = ws.cell(1,c,h); cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)
    for r,it in enumerate(items,2):
        vals = [it.get("proveedor",""),it.get("cotizacion",""),it.get("nro_item",""),it.get("codigo",""),it.get("marca",""),it.get("descripcion",""),it.get("grupo_comparable",""),it.get("moneda",""),it.get("cantidad_pedida",0),it.get("cantidad_real",0),it.get("unidad",""),it.get("precio_unitario",0),it.get("subtotal_sin_iva",0),it.get("iva_pct",0),it.get("iva_monto",0),it.get("total_con_iva",0),it.get("minimo_compra",""),it.get("entrega",""),it.get("notas",""),it.get("parser","")]
        for c,v in enumerate(vals,1):
            cell = ws.cell(r,c,v); cell.border = Border(left=thin,right=thin,top=thin,bottom=thin); cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c in [12,13,15,16]: cell.number_format = '$ #,##0.00'
            if c == 7 and not v: cell.fill = yellow
    for i,w in enumerate([20,18,8,18,16,60,24,10,14,14,10,14,16,10,14,16,20,20,42,16],1): ws.column_dimensions[get_column_letter(i)].width=w
    if items:
        tab = Table(displayName="ItemsDetectados", ref=f"A1:T{len(items)+1}"); tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True); ws.add_table(tab)
    ws2 = wb.create_sheet("Comparaciones"); headers2 = ["Grupo","Proveedor","Descripción","Moneda","Cant. real","P. unit.","Subtotal","IVA","Total","Dif. unit.","Dif. total","Estado","Nota"]
    for c,h in enumerate(headers2,1):
        cell = ws2.cell(1,c,h); cell.fill = fill; cell.font = font; cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)
    row = 2
    for comp in comps:
        for off in comp.get("ofertas", []):
            vals = [comp.get("grupo_comparable",""),off.get("proveedor",""),off.get("descripcion",""),off.get("moneda",""),off.get("cantidad_real",0),off.get("precio_unitario",0),off.get("subtotal_sin_iva",0),off.get("iva_monto",0),off.get("total_con_iva",0),off.get("dif_unit_vs_mejor",0),off.get("dif_total_vs_mejor",0),"MEJOR PRECIO" if off.get("recomendado_precio") else "Alternativa",comp.get("nota","")]
            for c,v in enumerate(vals,1):
                cell = ws2.cell(row,c,v); cell.border = Border(left=thin,right=thin,top=thin,bottom=thin); cell.alignment = Alignment(wrap_text=True, vertical="top")
                if c in [6,7,8,9,10,11]: cell.number_format = '$ #,##0.00'
                if c == 12 and v == "MEJOR PRECIO": cell.fill = green
            row += 1
    for i,w in enumerate([24,20,60,10,14,14,16,14,16,14,14,18,36],1): ws2.column_dimensions[get_column_letter(i)].width=w
    ws3 = wb.create_sheet("Pendientes de Revisión"); ws3["A1"]="Pendientes de Revisión"; ws3["A1"].font=Font(size=16,bold=True)
    headers3=["Proveedor","Descripción","Motivo","Grupo sugerido"]
    for c,h in enumerate(headers3,1):
        cell=ws3.cell(3,c,h); cell.fill=fill; cell.font=font
    r=4
    for it in [i for i in items if "universal" in i.get("parser","") or not i.get("grupo_comparable")]:
        ws3.cell(r,1,it.get("proveedor","")); ws3.cell(r,2,it.get("descripcion","")); ws3.cell(r,3,"Revisar extracción universal / grupo comparable"); ws3.cell(r,4,it.get("grupo_comparable","")); r+=1
    for i,w in enumerate([22,70,38,28],1): ws3.column_dimensions[get_column_letter(i)].width=w
    out=io.BytesIO(); wb.save(out); out.seek(0); return out

@app.post("/api/analyze")
async def analyze(files: List[UploadFile] = File(...)):
    all_items, raw = [], []
    for f in files:
        data = await f.read(); filename = f.filename or "archivo"
        if not filename.lower().endswith(".pdf"): continue
        try:
            pages = pdf_pages(data); text = "\n".join(pages); prov = detect_provider(filename, text)
            if prov == "Provemet / Complemet": items = parse_provemet(filename, pages)
            elif prov == "IVANAR": items = parse_ivanar(filename, pages)
            elif prov == "La Tornillera": items = parse_tornillera(filename, pages)
            elif prov == "HG Confecciones": items = parse_hg(filename, pages)
            elif prov == "Ingeniería Boggio": items = parse_boggio(filename, pages)
            elif prov == "Marlew": items = parse_marlew(filename, pages)
            elif prov == "Ateco": items = parse_ateco(filename, pages)
            else: items = parse_universal(filename, pages)
            if not items: items = parse_universal(filename, pages)
            all_items.extend(items); raw.append({"archivo": filename, "tabla": f"Texto PDF - {prov}", "columns": ["Texto"], "rows": [[line] for line in text.splitlines()[:600]]})
        except Exception as e:
            raw.append({"archivo": filename, "tabla": "ERROR", "columns": ["Error"], "rows": [[str(e)]]})
    return {"items": all_items, "raw_tables": raw, "warnings": ["V5: motor híbrido + lector universal + matching aproximado. Validar antes de emitir OC."]}

@app.post("/api/compare")
async def compare(payload: dict):
    comps, summary = compare_items(payload.get("items", []))
    return {"comparisons": comps, "summary": summary}

@app.post("/api/export_excel")
async def export_excel(payload: dict):
    out = build_excel(payload.get("items", []), payload.get("comparisons", []), payload.get("summary", ""))
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=comparativa_cotizaciones_v5.xlsx"})

@app.get("/api/health")
def health():
    return {"status": "ok", "version": "v5-industrial"}
